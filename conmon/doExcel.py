from openpyxl import  load_workbook

class DoExcel:
    # 打开excel文档
    def __init__(self,excelpath):
        wb= load_workbook(excelpath)
        self.sh_case_data=wb.get_sheet_by_name("apidata")



    def get_caseData(self):
        allCase=[]
        for index in range(2,self.sh_case_data.max_row+1):
            case_data={}
            case_data["case_id"] = self.sh_case_data.cell(row=index, column=1).value
            case_data["api_name"] = self.sh_case_data.cell(row=index, column=4).value
            case_data["method"] = self.sh_case_data.cell(row=index, column=5).value
            case_data["url"] = self.sh_case_data.cell(row=index, column=6).value
            case_data["request_data"] = self.sh_case_data.cell(row=index, column=7).value
            case_data["expected_data"] = self.sh_case_data.cell(row=index, column=8).value
            allCase.append(case_data)
        return allCase
